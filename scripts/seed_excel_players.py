import os
import openpyxl
import psycopg2

EXCEL_PATH = os.path.join("db", "players.xlsx")
DATA_CONTEXT_PATH = os.path.join("backend", "Data", "DataContext.cs")

def map_role(excel_role):
    role_map = {
        "Batter": "Batsman",
        "Batsman": "Batsman",
        "Bowler": "Bowler",
        "All-Rounder": "AllRounder",
        "AllRounder": "AllRounder",
        "Wicketkeeper Batter": "WicketKeeper",
        "Wicketkeeper-Batter": "WicketKeeper",
        "WicketKeeper": "WicketKeeper"
    }
    return role_map.get(excel_role, excel_role)

def load_players_from_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]
    
    players = []
    for r in range(2, sheet.max_row + 1):
        row = {headers[c]: sheet.cell(r, c + 1).value for c in range(len(headers))}
        
        raw_rating = float(row['rating']) if row['rating'] is not None else 0.0
        rating = round(raw_rating / 10.0, 1) if raw_rating > 10.0 else round(raw_rating, 1)
        base_price = float(row['base_price']) if row['base_price'] is not None else 20000000.0
        
        p = {
            "id": str(row['id']).strip(),
            "name": str(row['name']).strip(),
            "image_url": str(row['image_url']).strip() if row['image_url'] else "",
            "country": str(row['country']).strip() if row['country'] else "India",
            "role": map_role(str(row['role']).strip()),
            "batting_style": str(row['batting_style']).strip() if row['batting_style'] else "",
            "bowling_style": str(row['bowling_style']).strip() if row['bowling_style'] else "",
            "age": int(row['age']) if row['age'] is not None else 25,
            "rating": rating,
            "base_price": base_price,
            "category": str(row['category']).strip() if row['category'] else "Capped",
            "matches_played": int(row['matches played']) if row['matches played'] is not None else 0,
            "ipl_runs": int(row['ipl_runs']) if row['ipl_runs'] is not None else 0,
            "ipl_wickets": int(row['ipl_wickets']) if row['ipl_wickets'] is not None else 0,
            "strike_rate": float(row['strike_rate']) if row['strike_rate'] is not None else 0.0,
            "average": float(row['average']) if row['average'] is not None else 0.0,
            "fifties": int(row['50s']) if row['50s'] is not None else 0,
            "hundreds": int(row['100s']) if row['100s'] is not None else 0,
            "economy": float(row['economy']) if row['economy'] is not None else 0.0,
            "description": str(row['description']).strip() if row['description'] else ""
        }
        players.append(p)
    return players

def update_datacontext_cs(players):
    with open(DATA_CONTEXT_PATH, "r", encoding="utf-8") as f:
        content = f.read()
        
    csharp_seed_lines = []
    for p in players:
        esc_name = p['name'].replace('"', '\\"')
        esc_desc = p['description'].replace('"', '\\"')
        esc_bat = p['batting_style'].replace('"', '\\"')
        esc_bowl = p['bowling_style'].replace('"', '\\"')
        esc_country = p['country'].replace('"', '\\"')
        esc_img = p['image_url'].replace('"', '\\"')
        
        csharp_seed_lines.append(
            f'                    new Player {{ Id = Guid.Parse("{p["id"]}"), Name = "{esc_name}", ImageUrl = "{esc_img}", Country = "{esc_country}", Role = "{p["role"]}", BattingStyle = "{esc_bat}", BowlingStyle = "{esc_bowl}", Age = {p["age"]}, Rating = {p["rating"]}m, BasePrice = {p["base_price"]:.2f}m, Category = "{p["category"]}", MatchesPlayed = {p["matches_played"]}, IplRuns = {p["ipl_runs"]}, IplWickets = {p["ipl_wickets"]}, StrikeRate = {p["strike_rate"]}m, Average = {p["average"]}m, Fifties = {p["fifties"]}, Hundreds = {p["hundreds"]}, Economy = {p["economy"]}m, Description = "{esc_desc}" }}'
        )
        
    csharp_hasdata = "entity.HasData(\n" + ",\n".join(csharp_seed_lines) + "\n                );"
    
    start_tag = "entity.HasData("
    end_tag = ");"
    
    start_idx = content.find(start_tag)
    if start_idx == -1:
        print("Could not find entity.HasData in DataContext.cs")
        return
        
    end_idx = content.find(end_tag, start_idx)
    if end_idx == -1:
        print("Could not find closing ); for entity.HasData in DataContext.cs")
        return
        
    new_content = content[:start_idx] + csharp_hasdata + content[end_idx + len(end_tag):]
    
    with open(DATA_CONTEXT_PATH, "w", encoding="utf-8") as f:
        f.write(new_content)
    print(f"Updated {DATA_CONTEXT_PATH} with {len(players)} players.")

def execute_neon_db_update(players):
    conn_str = "Host=ep-rapid-lab-az8yvbrm-pooler.c-3.ap-southeast-1.aws.neon.tech;Database=neondb;Username=neondb_owner;Password=npg_duj5WX6peqnv;SSL Mode=Require;Trust Server Certificate=true;"
    params = dict(item.split('=', 1) for item in conn_str.rstrip(';').split(';') if '=' in item)
    dsn = f"host={params['Host']} dbname={params['Database']} user={params['Username']} password={params['Password']} sslmode=require"
    
    conn = psycopg2.connect(dsn)
    cur = conn.cursor()
    
    # Drop and recreate players table with exact column order
    cur.execute("DROP TABLE IF EXISTS players CASCADE;")
    cur.execute("""
    CREATE TABLE players (
        id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
        name VARCHAR(100) NOT NULL,
        image_url VARCHAR(255),
        country VARCHAR(100) NOT NULL DEFAULT 'India',
        role VARCHAR(50) NOT NULL,
        batting_style VARCHAR(50),
        bowling_style VARCHAR(50),
        age INT,
        rating DECIMAL(3, 1) NOT NULL DEFAULT 0.0,
        base_price DECIMAL(15, 2) NOT NULL DEFAULT 2000000.00,
        category VARCHAR(50) NOT NULL,
        matches_played INT DEFAULT 0,
        ipl_runs INT DEFAULT 0,
        ipl_wickets INT DEFAULT 0,
        strike_rate DECIMAL(5, 2) DEFAULT 0.0,
        average DECIMAL(5, 2) DEFAULT 0.0,
        fifties INT DEFAULT 0,
        hundreds INT DEFAULT 0,
        economy DECIMAL(4, 2) DEFAULT 0.0,
        description TEXT,
        created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP NOT NULL
    );
    CREATE INDEX IF NOT EXISTS idx_players_role ON players(role);
    CREATE INDEX IF NOT EXISTS idx_players_country ON players(country);
    
    ALTER TABLE rooms ADD CONSTRAINT fk_rooms_current_player FOREIGN KEY (current_player_id) REFERENCES players(id) ON DELETE SET NULL;
    ALTER TABLE bids ADD CONSTRAINT fk_bids_player FOREIGN KEY (player_id) REFERENCES players(id) ON DELETE CASCADE;
    ALTER TABLE purchases ADD CONSTRAINT fk_purchases_player FOREIGN KEY (player_id) REFERENCES players(id) ON DELETE CASCADE;
    """)
    conn.commit()
    print("Recreated players table in Neon PostgreSQL with updated columns.")
    
    insert_sql = """
    INSERT INTO players (id, name, image_url, country, role, batting_style, bowling_style, age, rating, base_price, category, matches_played, ipl_runs, ipl_wickets, strike_rate, average, fifties, hundreds, economy, description, created_at)
    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP);
    """
    
    rows_to_insert = [
        (
            p['id'], p['name'], p['image_url'], p['country'], p['role'],
            p['batting_style'], p['bowling_style'], p['age'], p['rating'],
            p['base_price'], p['category'], p['matches_played'], p['ipl_runs'],
            p['ipl_wickets'], p['strike_rate'], p['average'], p['fifties'],
            p['hundreds'], p['economy'], p['description']
        )
        for p in players
    ]
    
    cur.executemany(insert_sql, rows_to_insert)
    conn.commit()
    
    cur.execute("SELECT COUNT(*) FROM players;")
    count = cur.fetchone()[0]
    print(f"Successfully inserted {count} players into Neon PostgreSQL DB.")
    
    cur.close()
    conn.close()

if __name__ == "__main__":
    os.makedirs("scripts", exist_ok=True)
    players = load_players_from_excel()
    print(f"Loaded {len(players)} players from {EXCEL_PATH}.")
    
    update_datacontext_cs(players)
    execute_neon_db_update(players)
